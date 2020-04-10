# The following retrieves unique values from an existing dataset to form a column of investors in a new workbook

import openpyxl as xl
import xlsxwriter as xw

path = "Unicorns_Dataset.xlsx"

workbook = xl.load_workbook(path)
ws = workbook["Investors"]

new_workbook = xw.Workbook("Fintech Unicorn Investors.xlsx")
ws2 = new_workbook.add_worksheet()


def convert_columns_to_list(ws, column):
    new_list = []

    for i in range(1, ws.max_row + 1):
        company = ws.cell(row=i, column=column).value
        new_list.append(company)

    return new_list


def create_one_column_of_unique_values(list1, list2):
    for company in list2:
        if company in list1:
            continue
        else:
            list1.append(company)

    return list1


def add_list_to_worksheet(list, ws, row, column):
    for company in list:
        ws.write(row, column, company)
        row += 1


list1 = convert_columns_to_list(ws, 1)
list2 = convert_columns_to_list(ws, 2)
list3 = convert_columns_to_list(ws, 3)
list4 = convert_columns_to_list(ws, 4)

create_one_column_of_unique_values(list1, list2)
create_one_column_of_unique_values(list1, list3)
create_one_column_of_unique_values(list1, list4)

add_list_to_worksheet(list1, ws2, 0, 0)
new_workbook.close()
